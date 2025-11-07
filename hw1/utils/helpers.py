import os
import warnings
from datetime import datetime

import joblib
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")


def export_to_excel(df, queries=None):
    """
    Альтернативная версия экспорта в Excel
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"energy_analytics_export_{timestamp}.xlsx"
        filepath = os.path.join("exports", filename)

        os.makedirs("exports", exist_ok=True)

        # Создаем workbook вручную чтобы контролировать видимость листов
        workbook = Workbook()

        # Удаляем дефолтный лист если он есть
        if "Sheet" in workbook.sheetnames:
            std_sheet = workbook["Sheet"]
            workbook.remove(std_sheet)

        # 1. Основные данные (всегда видимый)
        ws1 = workbook.create_sheet("Исходные данные", 0)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)

        # 2. Сводная статистика
        if df is not None:
            stats_data = {
                "Метрика": ["Всего записей", "Период данных", "Средняя мощность"],
                "Значение": [
                    len(df),
                    f"{df['DateTime'].min().date()} - {df['DateTime'].max().date()}",
                    f"{df['Global_active_power'].mean():.2f}",
                ],
            }
            stats_df = pd.DataFrame(stats_data)

            ws2 = workbook.create_sheet("Сводная статистика", 1)
            for r in dataframe_to_rows(stats_df, index=False, header=True):
                ws2.append(r)

        # Сохраняем workbook
        workbook.save(filepath)
        return True, filename

    except Exception as e:
        return False, str(e)
