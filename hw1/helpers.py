import os
import warnings
from datetime import datetime

import joblib
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

warnings.filterwarnings("ignore")


def export_to_excel(df, queries=None):
    """
    Альтернативная версия экспорта в Excel
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"energy_analytics_export_{timestamp}.xlsx"
        filepath = os.path.join("exports", filename)

        os.makedirs("exports", exist_ok=True)

        # Создаем workbook вручную чтобы контролировать видимость листов
        workbook = Workbook()

        # Удаляем дефолтный лист если он есть
        if "Sheet" in workbook.sheetnames:
            std_sheet = workbook["Sheet"]
            workbook.remove(std_sheet)

        # 1. Основные данные (всегда видимый)
        ws1 = workbook.create_sheet("Исходные данные", 0)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws1.append(r)

        # 2. Сводная статистика
        if df is not None:
            stats_data = {
                "Метрика": ["Всего записей", "Период данных", "Средняя мощность"],
                "Значение": [
                    len(df),
                    f"{df['DateTime'].min().date()} - {df['DateTime'].max().date()}",
                    f"{df['Global_active_power'].mean():.2f}",
                ],
            }
            stats_df = pd.DataFrame(stats_data)

            ws2 = workbook.create_sheet("Сводная статистика", 1)
            for r in dataframe_to_rows(stats_df, index=False, header=True):
                ws2.append(r)

        # Сохраняем workbook
        workbook.save(filepath)
        return True, filename

    except Exception as e:
        return False, str(e)


# def create_sample_data(n_samples=1000):
#     """Создание примерных данных для тестирования"""
#     dates = pd.date_range(start="2006-01-01", end="2010-12-31", freq="H")[:n_samples]

#     data = {
#         "DateTime": dates,
#         "Global_active_power": np.random.exponential(1.5, n_samples),
#         "Global_reactive_power": np.random.exponential(0.2, n_samples),
#         "Voltage": np.random.normal(240, 5, n_samples),
#         "Global_intensity": np.random.exponential(6, n_samples),
#         "Sub_metering_1": np.random.exponential(0.5, n_samples),
#         "Sub_metering_2": np.random.exponential(0.3, n_samples),
#         "Sub_metering_3": np.random.exponential(2, n_samples),
#     }

#     df = pd.DataFrame(data)

#     # Добавляем временные признаки как в реальных данных
#     df = _add_temporal_features_to_sample(df)

#     return df


def _add_temporal_features_to_sample(df):
    """Добавление временных признаков к примерным данным"""
    df["Year"] = df["DateTime"].dt.year
    df["Month"] = df["DateTime"].dt.month
    df["Day"] = df["DateTime"].dt.day
    df["Hour"] = df["DateTime"].dt.hour
    df["DayOfWeek"] = df["DateTime"].dt.dayofweek
    df["Weekend"] = df["DayOfWeek"].isin([5, 6]).astype(int)
    df["Season"] = df["Month"].apply(_get_season)
    return df


def _get_season(month):
    """Определение сезона по месяцу"""
    if month in [12, 1, 2]:
        return "Winter"
    elif month in [3, 4, 5]:
        return "Spring"
    elif month in [6, 7, 8]:
        return "Summer"
    else:
        return "Autumn"
